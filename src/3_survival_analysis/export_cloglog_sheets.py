#!/usr/bin/env python3
"""
Export ClogLog (c-spec, age-bins + decade dummies) results to Excel.
One sheet per CP variable: EF, DefPact, Loot, Mountains, BDist, Territorial.
Rows = covariates; Columns = Signing / Victory / Fade × M1c–M6c.
"""
import sys, os, warnings
warnings.filterwarnings('ignore')

BASE = (
    '/Users/luciasauer/Library/CloudStorage/'
    'GoogleDrive-lucia.sauer@bse.eu/Mi unidad/'
    'EconAI/1_agreements_violence/agreements_violence'
)
sys.path.insert(0, os.path.join(BASE, 'src/3_survival_analysis/functions'))
os.chdir(os.path.join(BASE, 'src/3_survival_analysis'))

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from cloglog_hazard import fit_cloglog

# ── 1. Load and prepare data ──────────────────────────────────────────────────
print('Loading data...')
spell_q = pd.read_csv(os.path.join(BASE, 'data/output/conflict_level/spell_q.csv'))
spell_q['yq']     = pd.PeriodIndex(spell_q['yq'], freq='Q')
spell_q['year_q'] = spell_q['yq'].dt.year

spell_q['_last_q'] = (
    spell_q.groupby('conflict_id')['conflict_age_q'].transform('max')
    == spell_q['conflict_age_q']
)
spell_q['event_S'] = spell_q['is_first_agreement']
spell_q['event_V'] = (spell_q['_last_q'] & (spell_q['cause_label'] == 'not at risk')).astype(int)
spell_q['event_F'] = (spell_q['_last_q'] & (spell_q['cause_label'] == 'fade')).astype(int)

spell_q['d_2000s'] = (spell_q['year_q'] >= 2000).astype(int)
spell_q['d_2010s'] = (spell_q['year_q'] >= 2010).astype(int)
spell_q['d_2020s'] = (spell_q['year_q'] >= 2020).astype(int)

_bins_cut = [0, 4, 8, 16, 32, np.inf]
spell_q['age_bin'] = pd.cut(spell_q['conflict_age_q'], bins=_bins_cut,
                             labels=['1-4Q','5-8Q','9-16Q','17-32Q','33+Q'], right=True)
BIN_COLS = {'5-8Q': 'abin_5_8Q', '9-16Q': 'abin_9_16Q',
            '17-32Q': 'abin_17_32Q', '33+Q': 'abin_33pQ'}
for _lbl, _col in BIN_COLS.items():
    spell_q[_col] = (spell_q['age_bin'] == _lbl).astype(float)
age_bin_cols = list(BIN_COLS.values())
DEC_COLS     = ['d_2000s', 'd_2010s', 'd_2020s']

bin_x_ia_cols, bin_x_ef_cols = [], []
for b in age_bin_cols:
    spell_q[f'{b}_x_ia'] = spell_q[b] * spell_q['high_ia_bin']
    bin_x_ia_cols.append(f'{b}_x_ia')
    spell_q[f'{b}_x_ef'] = spell_q[b] * spell_q['high_fe_etfra_bin']
    bin_x_ef_cols.append(f'{b}_x_ef')

# ── 2. Tab definitions ────────────────────────────────────────────────────────
TABS = [
    ('high_fe_etfra_bin',           'EF',          'HighEF'),
    ('defpact_majpow_onset',        'DefPact',      '3P-Guar.'),
    ('loot_onset',                  'Loot',         'Lootable'),
    ('mountains_onset',             'Mountains',    'Mountains'),
    ('bdist_onset',                 'BDist',        'BDist'),
    ('territorial_incompatibility', 'Territorial',  'Territorial'),
]

# Compute CP interaction columns
for cp, _, _ in TABS:
    if cp != 'high_fe_etfra_bin':
        spell_q[f'log_age_x_{cp}'] = spell_q['log_conflict_age_q'] * spell_q[cp]
        for b in age_bin_cols:
            col = f'{b}_x_{cp}'
            if col not in spell_q.columns:
                spell_q[col] = spell_q[b] * spell_q[cp]

# ── 3. Spec builder ───────────────────────────────────────────────────────────
def make_specs(cp):
    bx_cp = bin_x_ef_cols if cp == 'high_fe_etfra_bin' else [f'{b}_x_{cp}' for b in age_bin_cols]
    return {
        'M1c': age_bin_cols + DEC_COLS,
        'M2c': age_bin_cols + DEC_COLS + ['high_ia_bin'],
        'M3c': age_bin_cols + DEC_COLS + ['high_ia_bin'] + bin_x_ia_cols,
        'M4c': age_bin_cols + DEC_COLS + [cp],
        'M5c': age_bin_cols + DEC_COLS + [cp] + bx_cp,
        'M6c': age_bin_cols + DEC_COLS + ['high_ia_bin'] + bin_x_ia_cols + [cp] + bx_cp,
    }

# ── 4. Variable labels ────────────────────────────────────────────────────────
def make_labels(cp, cp_lbl):
    d = {
        'abin_5_8Q':   '5–8 Q',
        'abin_9_16Q':  '9–16 Q',
        'abin_17_32Q': '17–32 Q',
        'abin_33pQ':   '≥33 Q',
        'd_2000s':     '≥2000',
        'd_2010s':     '≥2010',
        'd_2020s':     '≥2020',
        'high_ia_bin': 'HighIA',
        cp:            cp_lbl,
    }
    for b in age_bin_cols:
        b_s = b.replace('abin_', '').replace('pQ', '+Q')
        d[f'{b}_x_ia']   = f'{b_s} × HighIA'
        d[f'{b}_x_ef']   = f'{b_s} × HighEF'
        d[f'{b}_x_{cp}'] = f'{b_s} × {cp_lbl}'
    return d

# ── 5. Fit helper ─────────────────────────────────────────────────────────────
OUTCOMES   = [('event_S', 'Signing'), ('event_V', 'Victory'), ('event_F', 'Fade')]
SPEC_ORDER = ['M1c', 'M2c', 'M3c', 'M4c', 'M5c', 'M6c']

def fit_all(cp):
    specs = make_specs(cp)
    fits  = {}
    for outcome, _ in OUTCOMES:
        fits[outcome] = {}
        for spec in SPEC_ORDER:
            covs   = specs[spec]
            df_use = spell_q.dropna(subset=covs)
            try:
                tbl, _ = fit_cloglog(outcome, covs, df_use)
                fits[outcome][spec] = tbl
            except Exception as e:
                print(f'    WARN {outcome} {spec}: {e}')
                fits[outcome][spec] = None
    return specs, fits

def get_val(tbl, cov):
    """Return (hr_str, se_str) for a covariate in a result table."""
    if tbl is None or cov not in tbl.index:
        return '—', ''
    hr  = tbl.loc[cov, 'HR']
    se  = tbl.loc[cov, 'se']
    sig = str(tbl.loc[cov, 'sig']).strip()
    return f'{hr:.3f}{sig}', f'({se:.3f})'

# ── 6. Excel styles ───────────────────────────────────────────────────────────
C_SIGN_DARK  = '7B1518'
C_SIGN_MED   = 'C0392B'
C_SIGN_LIGHT = 'FADBD8'
C_VICT_DARK  = '7B3200'
C_VICT_MED   = 'C05C00'
C_VICT_LIGHT = 'FEF0E7'
C_FADE_DARK  = '1A3A5C' 
C_FADE_MED   = '2B6CB0' 
C_FADE_LIGHT = 'EBF2FA'
C_GRAY       = 'F2F2F2'
C_GRAY2      = 'D9D9D9'

OUT_COLORS = {
    'Signing': (C_SIGN_DARK, C_SIGN_MED, C_SIGN_LIGHT),
    'Victory': (C_VICT_DARK, C_VICT_MED, C_VICT_LIGHT),
    'Fade':    (C_FADE_DARK, C_FADE_MED, C_FADE_LIGHT),
}

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(bold=False, italic=False, size=10, color='000000'):
    return Font(bold=bold, italic=italic, size=size, color=color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center')

thin_side   = Side(style='thin',   color='AAAAAA')
medium_side = Side(style='medium', color='888888')
thin_border = Border(bottom=thin_side)

def right_border():
    return Border(right=Side(style='medium', color='888888'))

# ── 7. Build each sheet ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

for cp, tab_name, cp_lbl in TABS:
    print(f'  Building sheet: {tab_name} ({cp})')
    specs, fits = fit_all(cp)
    labels      = make_labels(cp, cp_lbl)

    # Ordered covariate list
    all_covs = []
    for spec in SPEC_ORDER:
        for c in specs[spec]:
            if c not in all_covs:
                all_covs.append(c)

    ws = wb.create_sheet(title=tab_name)
    ws.sheet_view.showGridLines = False

    n_specs = len(SPEC_ORDER)   # 6
    n_out   = len(OUTCOMES)     # 3
    total_data_cols = n_out * n_specs  # 18

    # ── Row 1: outcome group headers ──────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.cell(1, 1, '').fill = fill(C_GRAY2)
    col = 2
    for out_label, (dark, med, light) in zip(
            [o for _, o in OUTCOMES], OUT_COLORS.values()):
        c = ws.cell(1, col, out_label)
        c.font      = font(bold=True, size=11, color='FFFFFF')
        c.fill      = fill(dark)
        c.alignment = center()
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1,   end_column=col + n_specs - 1)
        # Right border after each outcome group
        for r in range(1, 1 + 2 + 2 * len(all_covs) + 5):
            ws.cell(r, col + n_specs - 1).border = right_border()
        col += n_specs

    # ── Row 2: spec sub-headers ───────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.cell(2, 1, 'Variable').font      = font(bold=True, size=10, color='FFFFFF')
    ws.cell(2, 1).fill      = fill('444444')
    ws.cell(2, 1).alignment = left()
    col = 2
    for _, out_label in OUTCOMES:
        dark, med, light = OUT_COLORS[out_label]
        for spec in SPEC_ORDER:
            c = ws.cell(2, col, spec)
            c.font      = font(bold=True, size=9, color='FFFFFF')
            c.fill      = fill(med)
            c.alignment = center()
            col += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 3
    for cov in all_covs:
        lbl      = labels.get(cov, cov)
        is_bin   = cov in age_bin_cols
        is_dec   = cov in DEC_COLS
        is_ia    = cov == 'high_ia_bin'
        is_cp    = cov == cp
        is_inter = '_x_' in cov
        bg = C_GRAY if (is_bin or is_dec) else ('FFFDE7' if is_ia else
             ('E8F5E9' if is_cp else ('FAFAFA' if is_inter else 'FFFFFF')))

        ws.row_dimensions[row].height     = 16
        ws.row_dimensions[row + 1].height = 14

        # Variable label (spans 2 rows)
        label_cell = ws.cell(row, 1, lbl)
        label_cell.font      = font(bold=not is_inter, size=9 if is_inter else 10,
                                    color='444444' if (is_bin or is_dec) else '000000')
        label_cell.fill      = fill(bg)
        label_cell.alignment = left()
        ws.cell(row + 1, 1).fill = fill(bg)

        col = 2
        for outcome, out_lbl in OUTCOMES:
            dark, med, light = OUT_COLORS[out_lbl]
            for spec in SPEC_ORDER:
                hr_str, se_str = get_val(fits[outcome][spec], cov)
                # HR cell
                hc = ws.cell(row, col, hr_str)
                hc.font      = font(bold=(hr_str != '—' and any(s in hr_str for s in ['*'])),
                                    size=10)
                hc.fill      = fill(light if hr_str != '—' else 'F8F8F8')
                hc.alignment = center()
                # SE cell
                sc = ws.cell(row + 1, col, se_str)
                sc.font      = font(size=9, color='555555')
                sc.fill      = fill(light if se_str else 'F8F8F8')
                sc.alignment = center()
                col += 1
        row += 2

    # ── Separator blank row ───────────────────────────────────────────────────
    ws.row_dimensions[row].height = 8
    for c in range(1, 2 + total_data_cols):
        ws.cell(row, c).fill = fill('DDDDDD')
    row += 1

    # ── AIC / N obs / Events rows ─────────────────────────────────────────────
    for stat_label in ('AIC', 'N obs', 'Events'):
        key_map = {'AIC': 'aic', 'N obs': 'n_obs', 'Events': 'n_events'}
        ws.row_dimensions[row].height = 15
        c0 = ws.cell(row, 1, stat_label)
        c0.font      = font(italic=True, size=9, color='555555')
        c0.fill      = fill(C_GRAY)
        c0.alignment = left()
        col = 2
        for outcome, _ in OUTCOMES:
            for spec in SPEC_ORDER:
                tbl = fits[outcome][spec]
                if tbl is not None:
                    raw = tbl[key_map[stat_label]].iloc[0]
                    val = f'{int(raw):,}' if stat_label in ('N obs', 'Events') else f'{raw:.1f}'
                else:
                    val = '—'
                c = ws.cell(row, col, val)
                c.font      = font(italic=True, size=9, color='555555')
                c.fill      = fill(C_GRAY)
                c.alignment = center()
                col += 1
        row += 1

    # ── Note row ──────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    note = ws.cell(row, 1, '*** p<0.01  ** p<0.05  * p<0.10  |  HR from ClogLog (age-bins + decade dummies)  |  SE of log-HR in parentheses')
    note.font      = font(italic=True, size=8, color='777777')
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=1 + total_data_cols)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 24
    for col_idx in range(2, 2 + total_data_cols):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    ws.freeze_panes = 'B3'

# ── 8. Save ───────────────────────────────────────────────────────────────────
out_path = os.path.join(BASE, 'src/3_survival_analysis/results/cloglog_c_specs.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'\nSaved: {out_path}')
